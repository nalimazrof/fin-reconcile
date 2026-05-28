import os
import glob
import pandas as pd
import re
import time
import requests
import base64
import warnings
from datetime import datetime
import openpyxl

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# ==========================================
# 1. 基础与路径配置 (已升级为完全动态流转)
# ==========================================
USER_HOME = os.path.expanduser("~")
EXCEL_FILE = os.path.join(USER_HOME, "OneDrive", "Finances.xlsx")
CSV_DIR = "CSV for validation"
ASSETS_TAB = "US assets"
MAX_DAYS_TOLERANCE = 5 
TOKEN_FILE = ".simplefin_url"

# 🚀 【升级二】：让 YEAR_TAB 根据当前系统时间全自动更替 (例如 2026年自动变为 "2026")
YEAR_TAB = str(datetime.now().year)

def get_date_range():
    print(f"\n📅 当前系统年份：{YEAR_TAB} 年")
    print("请输入需要核对的时间范围 (格式: YYYY-MM-DD)")
    start_str = input("请输入开始日期 (例如 2026-04-15): ").strip()
    end_str = input("请输入结束日期 (例如 2026-05-15): ").strip()
    
    try:
        start_date = pd.to_datetime(start_str)
        end_date = pd.to_datetime(end_str)
        return start_date, end_date
    except Exception as e:
        print("❌ 日期格式输入有误，请按 YYYY-MM-DD 格式重新运行。")
        exit()

def get_or_claim_access_url():
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, 'r') as f:
            return f.read().strip()

    print("\n⚠️ 未在本地找到 API 凭证。")
    setup_token = input("👉 请粘贴在 SimpleFIN 生成的 Setup Token (直接回车则跳过并使用纯CSV模式): ").strip()
    
    if not setup_token:
        return None

    try:
        claim_url = base64.b64decode(setup_token).decode('utf-8')
        print("🔄 正在向 SimpleFIN 兑换永久 Access URL...")
        response = requests.post(claim_url, timeout=10)
        response.raise_for_status()
        access_url = response.text.strip()
        with open(TOKEN_FILE, 'w') as f:
            f.write(access_url)
        print("✅ 兑换成功！永久凭证已安全保存在本目录的 .simplefin_url 文件中。")
        return access_url
    except Exception as e:
        print(f"❌ Setup Token 兑换失败: {e}")
        return None

def fetch_simplefin_data(start_date, end_date):
    access_url = get_or_claim_access_url()
    if not access_url:
        print("⚠️ 降级提示：将仅扫描本地 CSV 文件夹。")
        return []

    print("\n🌐 正在连接 SimpleFIN 网关...")
    try:
        scheme, rest = access_url.split('//', 1)
        auth, rest = rest.split('@', 1)
        url = scheme + '//' + rest + '/accounts'
        username, password = auth.split(':', 1)

        start_ts = int(time.mktime(start_date.timetuple()))
        end_ts = int(time.mktime(end_date.timetuple()))

        response = requests.get(url, auth=(username, password), params={"start-date": start_ts, "end-date": end_ts})
        response.raise_for_status()
        data = response.json()
        print(f"✅ 成功连接云端！共获取 {len(data.get('accounts', []))} 个账户数据。")
        return data.get('accounts', [])
    except Exception as e:
        print(f"❌ API 连接失败 ({e})。\n🔄 已自动降级为纯本地 CSV 备用模式。")
        return []

def extract_last_4_digits(val):
    if pd.isna(val) or not val: return ""
    s = str(val).split('.')[0] 
    s = re.sub(r'\D', '', s)
    return s[-4:] if len(s) >= 4 else ""

def build_account_mapping(excel_path):
    try:
        df_assets = pd.read_excel(excel_path, sheet_name=ASSETS_TAB)
    except Exception:
        print(f"❌ 找不到 '{ASSETS_TAB}' 标签页，无法建立自动映射。")
        return None, None

    account_map = {}
    all_account_names = []

    for _, row in df_assets.iterrows():
        name = str(row.get('Name', '')).strip()
        if not name or name == 'nan': continue
            
        all_account_names.append(name)
        acct_type = str(row.get('Type', '')).strip().lower()
        card_4 = extract_last_4_digits(row.get('Card#'))
        acc_4 = extract_last_4_digits(row.get('Account#'))

        suffix = card_4 if acct_type == 'credit' and card_4 else acc_4 if acct_type != 'credit' and acc_4 else ""
        if suffix:
            account_map[suffix] = name
            
    all_account_names.sort(key=len, reverse=True)
    return all_account_names, account_map

def standardize_bank_csv(csv_path):
    df = pd.read_csv(csv_path, index_col=False)
    df.columns = df.columns.str.strip()

    date_col = next((c for c in ['Posting Date', 'Trans. Date', 'Transaction Date', 'Date'] if c in df.columns), None)
    amount_col = next((c for c in ['Amount', 'Transaction Amount', 'Debit', 'Credit'] if c in df.columns), None)

    if not date_col or not amount_col: return None
    desc_col = next((c for c in ['Description', 'Payee', 'Details'] if c in df.columns), df.columns[0])

    df = df.rename(columns={date_col: 'Date', amount_col: 'Amount', desc_col: 'Description'})
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    
    df['Amount'] = df['Amount'].astype(str).str.replace(r'[^\d\.\-]', '', regex=True)
    df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
    df = df.dropna(subset=['Date', 'Amount'])
    return df[['Date', 'Amount', 'Description']]

def append_transactions_to_excel(excel_path, sheet_name, transactions_to_add):
    """核心函数：将新交易追加到 Excel 中，并智能克隆上一行的日期格式和样式"""
    if not transactions_to_add:
        return
    
    print(f"\n💾 正在打开 Excel 账本准备写入...")
    wb = openpyxl.load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        print(f"❌ 找不到标签页 {sheet_name}，放弃写入。")
        return
        
    ws = wb[sheet_name]
    
    # 1. 建立表头索引并寻找 Label 列
    headers = [str(cell.value).strip() for cell in ws[1]]
    try:
        idx_date = headers.index('Date') + 1
        idx_amount = headers.index('Amount') + 1
        idx_from = headers.index('from Account') + 1
        idx_to = headers.index('to Account') + 1
        idx_comments = headers.index('Comments') + 1
    except ValueError as e:
        print(f"❌ Excel 表头结构不匹配，无法自动追记。缺失列: {e}")
        return

    if 'Label' in headers:
        idx_label = headers.index('Label') + 1
    else:
        idx_label = 6
        ws.cell(row=1, column=idx_label, value='Label')
        print("ℹ️ 未在表头发现 'Label' 列，已自动创建。")

    # 精准探测真正的非空最后一行
    real_max_row = ws.max_row
    for row_idx in range(ws.max_row, 0, -1):
        if any(ws.cell(row=row_idx, column=col_idx).value is not None for col_idx in range(1, len(headers) + 1)):
            real_max_row = row_idx
            break

    start_row = real_max_row + 1
    source_row = real_max_row  # 样式与格式克隆源

    # 🚀【核心优化】：动态获取上一行日期单元格的数字格式（例如 'M/D/YYYY'）
    # 如果上一行没有指定格式，则默认使用通用的 'YYYY-MM-DD' 作为兜底
    source_date_format = 'YYYY-MM-DD'
    if source_row >= 2:
        src_date_cell = ws.cell(row=source_row, column=idx_date)
        if src_date_cell.number_format:
            source_date_format = src_date_cell.number_format

    added_count = 0

    # 2. 逐行写入数据并同步克隆样式
    for tx in transactions_to_add:
        current_row = start_row + added_count
        
        # 写入真正的 datetime.date 对象（确保底层数字是 46155）
        target_date_obj = tx['Date'].to_pydatetime().date()
        
        cell_date = ws.cell(row=current_row, column=idx_date, value=target_date_obj)
        cell_amount = ws.cell(row=current_row, column=idx_amount, value=abs(tx['Amount']))
        cell_from = ws.cell(row=current_row, column=idx_from, value=tx['from_acc'])
        cell_to = ws.cell(row=current_row, column=idx_to, value=tx['to_acc'])
        cell_comments = ws.cell(row=current_row, column=idx_comments, value=tx['Description'])
        cell_label = ws.cell(row=current_row, column=idx_label, value="Auto-Append")
        
        # 🚀【核心修复】：让新行直接继承上一行的日期格式，不再硬编码
        cell_date.number_format = source_date_format
        
        # 🎨 克隆其余字体、背景斑马纹、边框、对齐方式等样式
        if source_row >= 2:
            for idx in [idx_date, idx_amount, idx_from, idx_to, idx_comments, idx_label]:
                src_cell = ws.cell(row=source_row, column=idx)
                dst_cell = ws.cell(row=current_row, column=idx)
                
                if src_cell.has_style:
                    if src_cell.font: dst_cell.font = openpyxl.styles.Font(**src_cell.font.__dict__)
                    if src_cell.border: dst_cell.border = openpyxl.styles.Border(**src_cell.border.__dict__)
                    if src_cell.fill: dst_cell.fill = openpyxl.styles.PatternFill(**src_cell.fill.__dict__)
                    if src_cell.alignment: dst_cell.alignment = openpyxl.styles.Alignment(**src_cell.alignment.__dict__)
                    # 金额等非日期列继续沿用各自原有的数字格式
                    if idx != idx_date and src_cell.number_format: 
                        dst_cell.number_format = src_cell.number_format

        added_count += 1

    final_row = start_row + added_count - 1

    # 3. 强制把 Excel Table 范围撑大
    if ws.tables:
        for table_name in list(ws.tables.keys()):
            table_obj = ws.tables[table_name]
            if isinstance(table_obj, str):
                continue
            try:
                old_ref = table_obj.ref
                match = re.match(r"^([A-Z]+[0-9]+):([A-Z]+)([0-9]+)$", old_ref)
                if match:
                    start_part, end_col, _ = match.groups()
                    new_ref = f"{start_part}:{end_col}{final_row}"
                    table_obj.ref = new_ref
                    print(f"📐 已成功将 Excel Table 【{table_name}】 边界扩充至: {new_ref}")
            except AttributeError:
                pass

    wb.save(excel_path)
    print(f"🎉 终极解脱！已成功追加 {added_count} 笔账目。")

def run_reconciliation():
    print("="*60)
    print("🚀 个人财务全自动对账系统 (双轨混合+智能Table扩充版)")
    print("="*60)

    print(f"📊 目标账本: {EXCEL_FILE}")
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ 找不到 Excel 账本，请检查路径。")
        return

    # 在正式运行前，检测今年对应的标签页是否存在
    wb_check = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    if YEAR_TAB not in wb_check.sheetnames:
        print(f"❌ 报错：您的 Excel 账本中还没有建名为 '{YEAR_TAB}' 的标签页！")
        print(f"💡 提示：请先在 Excel 里为今年新建一个 Tab 页，然后重新运行。")
        return
    wb_check.close()

    print(f"🔄 正在读取 Excel 建立账户尾号指纹库...")
    all_account_names, account_map = build_account_mapping(EXCEL_FILE)
    if not all_account_names: return

    start_date, end_date = get_date_range()
    bank_data_pool = {}

    # --- 阶段 A：填充 API 数据 ---
    sf_accounts = fetch_simplefin_data(start_date, end_date)
    for acc_data in sf_accounts:
        full_sf_name = f"{acc_data.get('org', {}).get('name', '')} {acc_data.get('name', '')}"
        matched_account = None

        for acc_name in all_account_names:
            if acc_name.lower() in full_sf_name.lower():
                matched_account = acc_name
                break
        if not matched_account:
            digits_groups = re.findall(r'\d+', acc_data.get('name', ''))
            if digits_groups:
                for suffix, acc_name in account_map.items():
                    if digits_groups[-1].endswith(suffix):
                        matched_account = acc_name
                        break
        
        if matched_account and acc_data.get('transactions'):
            df_api = pd.DataFrame(acc_data['transactions'])
            df_api['Date'] = pd.to_datetime(df_api['posted'], unit='s')
            df_api['Raw_Amount'] = pd.to_numeric(df_api['amount'], errors='coerce')
            df_api['Amount'] = df_api['Raw_Amount'].abs()
            df_api['Description'] = df_api['payee'].fillna('') + " " + df_api['description'].fillna('')
            
            mask = (df_api['Date'] >= start_date) & (df_api['Date'] <= end_date)
            bank_data_pool[matched_account] = { 'source': 'API', 'df': df_api.loc[mask] }

    # --- 阶段 B：扫描本地 CSV 并覆盖 ---
    if not os.path.exists(CSV_DIR):
        os.makedirs(CSV_DIR)
        
    csv_files = glob.glob(os.path.join(CSV_DIR, "*.csv"))
    for csv_file in csv_files:
        filename = os.path.basename(csv_file)
        matched_account = None

        for acc_name in all_account_names:
            if acc_name.lower() in filename.lower():
                matched_account = acc_name
                break
        if not matched_account:
            digits_groups = re.findall(r'\d+', filename)
            if digits_groups:
                for suffix, acc_name in account_map.items():
                    if digits_groups[0].endswith(suffix):
                        matched_account = acc_name
                        break

        if matched_account:
            df_csv = standardize_bank_csv(csv_file)
            if df_csv is not None and not df_csv.empty:
                df_csv['Raw_Amount'] = df_csv['Amount']
                df_csv['Amount'] = df_csv['Amount'].abs()
                mask = (df_csv['Date'] >= start_date) & (df_csv['Date'] <= end_date)
                bank_data_pool[matched_account] = { 'source': '本地 CSV', 'df': df_csv.loc[mask] }

    # --- 阶段 C：执行核对 ---
    if not bank_data_pool:
        print("\n⚠️ 没有获取到任何有效数据。")
        return

    df_excel = pd.read_excel(EXCEL_FILE, sheet_name=YEAR_TAB)
    df_excel['Date'] = pd.to_datetime(df_excel['Date'])
    df_excel['Amount'] = pd.to_numeric(df_excel['Amount'], errors='coerce').abs()
    df_excel['from Account'] = df_excel['from Account'].fillna('')
    df_excel['to Account'] = df_excel['to Account'].fillna('')

    all_pending_writebacks = []

    for matched_account, data_info in bank_data_pool.items():
        source_label = data_info['source']
        df_bank = data_info['df']
        
        print("-" * 60)
        print(f"⚙️ 核对账户: 【{matched_account}】 (数据源: {source_label})")
        
        excel_start = start_date - pd.Timedelta(days=MAX_DAYS_TOLERANCE)
        excel_end = end_date + pd.Timedelta(days=MAX_DAYS_TOLERANCE)
        
        excel_mask = (
            ((df_excel['from Account'] == matched_account) | (df_excel['to Account'] == matched_account)) &
            (df_excel['Date'] >= excel_start) & (df_excel['Date'] <= excel_end)
        )
        df_excel_acc = df_excel.loc[excel_mask].copy()
        
        bank_records = df_bank.to_dict('records')
        excel_records = df_excel_acc.to_dict('records')
        
        matched_e_idx, matched_b_idx = set(), set()

        for b_idx, b_row in enumerate(bank_records):
            for e_idx, e_row in enumerate(excel_records):
                if e_idx in matched_e_idx: continue 
                
                if abs(b_row['Amount'] - e_row['Amount']) < 0.01:
                    if abs((b_row['Date'] - e_row['Date']).days) <= MAX_DAYS_TOLERANCE:
                        matched_e_idx.add(e_idx)
                        matched_b_idx.add(b_idx)
                        break

        excel_only = [e for i, e in enumerate(excel_records) if i not in matched_e_idx and start_date <= e['Date'] <= end_date]
        bank_only = [b for i, b in enumerate(bank_records) if i not in matched_b_idx]

        if not excel_only and not bank_only:
            print("  ✅ 完美匹配！")
        else:
            if excel_only:
                df_e = pd.DataFrame(excel_only)
                df_e['Date'] = df_e['Date'].dt.date
                print("\n  ❌ 异常 A: Excel 有，银行没找到：")
                print(df_e[['Date', 'Amount', 'Comments']].to_string(index=False))
            if bank_only:
                df_b = pd.DataFrame(bank_only)
                print("\n  ❌ 异常 B: 银行有，Excel 没记（漏账）：")
                
                df_b_print = df_b.copy()
                df_b_print['Date'] = df_b_print['Date'].dt.date
                print(df_b_print[['Date', 'Amount', 'Description']].to_string(index=False))

                for b_row in bank_only:
                    is_income = b_row.get('Raw_Amount', -1) > 0 if data_info['source'] == 'API' else False
                    all_pending_writebacks.append({
                        'Date': b_row['Date'],
                        'Amount': b_row['Amount'],
                        'Description': b_row['Description'],
                        'from_acc': '' if is_income else matched_account,
                        'to_acc': matched_account if is_income else ''
                    })

    # --- 阶段 D：智能反向追记交互 ---
    if all_pending_writebacks:
        print("\n" + "="*60)
        print(f"🛠️  反向追记引擎触发：共发现 {len(all_pending_writebacks)} 笔未记账目！")
        print("="*60)
        
        user_choice = input("👉 是否要将这些漏记流水自动补录进 Excel？\n(输入 Y：逐笔确认 | 输入 ALL：一键全部补录 | 输入 N：跳过): ").strip().upper()
        
        to_write = []
        if user_choice == 'ALL':
            to_write = all_pending_writebacks
        elif user_choice == 'Y':
            for tx in all_pending_writebacks:
                direction = f"→ 入账到 【{tx['to_acc']}】" if tx['to_acc'] else f"← 从 【{tx['from_acc']}】 支出"
                print(f"\n[待补录] {tx['Date'].strftime('%Y-%m-%d')} | ${tx['Amount']:.2f} | {direction}\n描述: {tx['Description']}")
                confirm = input("是否补录该笔交易？(Y/N): ").strip().upper()
                if confirm == 'Y':
                    to_write.append(tx)
        
        if to_write:
            append_transactions_to_excel(EXCEL_FILE, YEAR_TAB, to_write)
        else:
            print("\n☕ 操作已取消，未对 Excel 账本进行任何修改。")

if __name__ == "__main__":
    run_reconciliation()